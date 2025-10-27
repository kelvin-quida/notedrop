from django.db import connection
from django.core.management.color import no_style
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from rest_framework import generics, status
from rest_framework.response import Response

from .models import Message
from .serializers import MessageSerializer

class MessageListCreateView(generics.ListCreateAPIView):
    queryset = Message.objects.all()
    serializer_class = MessageSerializer

    def delete(self, request, *args, **kwargs):
        count = self.get_queryset().count()
        self.get_queryset().delete()

        sequence_sql = connection.ops.sequence_reset_sql(no_style(), [Message])
        if sequence_sql:
            with connection.cursor() as cursor:
                for sql in sequence_sql:
                    cursor.execute(sql)

        return Response(
            {'message': f'{count} mensagens foram apagadas com sucesso.'},
            status=status.HTTP_200_OK
        )

class MessageUpdateView(generics.UpdateAPIView):
    queryset = Message.objects.all()
    serializer_class = MessageSerializer

class MessageDeleteView(generics.DestroyAPIView):
    queryset = Message.objects.all()
    serializer_class = MessageSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        message_id = instance.id
        instance.delete()
        return Response(
            {'message': f'Mensagem #{message_id} foi apagada com sucesso.'},
            status=status.HTTP_200_OK
        )


class MessageExportView(generics.GenericAPIView):
    queryset = Message.objects.all().order_by('-created_at')

    def get(self, request, *args, **kwargs):
        messages = self.get_queryset()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Messages"

        headers = ['ID', 'Content', 'Created At', 'Updated At']
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row, msg in enumerate(messages, start=2):
            worksheet.cell(row=row, column=1, value=msg.id)
            worksheet.cell(row=row, column=2, value=msg.content)
            worksheet.cell(
                row=row, column=3,
                value=msg.created_at.strftime('%Y-%m-%d %H:%M:%S')
            )
            worksheet.cell(
                row=row, column=4,
                value=getattr(msg, 'updated_at', msg.created_at).strftime('%Y-%m-%d %H:%M:%S')
            )

        for column in worksheet.columns:
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            column_letter = column[0].column_letter
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="messages_export.xlsx"'
        workbook.save(response)
        return response